#!/usr/bin/env python

import sqlite3
from openpyxl import load_workbook
# maybe after use this
# from pysqlcipher3 import dbapi2 as sqlite
import pystache
import os
import datetime
import argparse
import shutil

# if you have a encrypt tools, do that!
# import base64


# data model
class Attribute(object):
    def __init__(self, name: str):
        self._name = name

    def property(self):
        return self._name


class ClassMember(object):
    def __init__(self, attributes: Attribute, data_type: str, name: str):
        self._attributes = attributes
        self._type = data_type
        self._name = name

    def type_name(self):
        return self._type

    def var_name(self):
        return self._name

    def attributes(self):
        return self._attributes


class ClassDeclare(object):
    def __init__(self, name, date, lst):
        self._name = name
        self._classMembers = lst
        self._date = date

    def name(self):
        return self._name

    def class_members(self):
        return self._classMembers

    def date(self):
        return self._date


def info(msg):
    print("info: {}".format(msg))


def get_class_declare(class_name, primary_index, types, names):
    lst = []
    for i in range(0, len(types)):
        attributes = []
        if i == primary_index:
            attributes.append(Attribute("PrimaryKey"))
        item = ClassMember(attributes, types[i], names[i])
        if item is not None:
            lst.append(item)

    ret = ClassDeclare(class_name, datetime.datetime.now(), lst)
    return ret


def set_parser():
    parser = argparse.ArgumentParser()
    parser.add_argument("-o", "--output", help="db output path and name", required=True)
    parser.add_argument("-p", "--cspath", help="path of exported csharp files", required=True)
    parser.add_argument("-e", "--excel", help="path of xlsx", required=True)
    args = parser.parse_args()
    if not args.output or not args.excel:
        print("arguments is wrong!")
        exit(1)

    return args


def create_connection(db_path):
    """ Create a database connection to the SQLite database
    :param db_path: Database file path
    :return: Connection Object (or None)
    """
    try:
        conn = sqlite3.connect(db_path)
        return conn
    except sqlite3.Error as e:
        print(e)

    return None


def execute_query(conn, query):
    """

    :param conn: Connection object
    :param query: CREATE TABLE statement
    :return:
    """
    try:
        c = conn.cursor()
        c.execute(query)
    except sqlite3.Error as e:
        print(e)
        exit(1)


def main():
    args = set_parser()

    wb = load_workbook(filename=args.excel, data_only=True)

    if not os.path.exists(args.cspath):
        os.mkdir(args.cspath)

    # load mustache template
    class_template = ""
    with open("class.mustache") as class_file:
        class_template = class_file.read()

    if class_template == "":
        print("class template is empty!")
        exit(1)

    # check db exist
    if os.path.exists("master.db"):
        os.remove("master.db")

    # renew db file
    conn = create_connection("master.db")
    info("created db file")
    cursor = conn.cursor()

    # load from xlsx
    sheet_count = 0
    sheet_total_count = len(wb.sheetnames)
    for sheet in wb:
        sheet_count = sheet_count + 1
        # skip keyword
        if sheet.title[0] == '_':
            info("skipped sheet({}/{}) - {}".format(sheet_count, sheet_total_count, sheet.title))
            continue

        info('proceed sheet({}/{}) - {}'.format(sheet_count, sheet_total_count, sheet.title))
        field_name_list = list()
        field_type_list = list()

        row_index = 0
        primary_index = 0
        q = ""
        for cell in sheet:
            col_index = 0
            empty_check = 0

            if row_index >= 5:
                q = "insert into {} values (".format(sheet.title)

            for data in cell:
                if row_index == 2:
                    if data.value == "PrimaryKey":
                        primary_index = col_index
                        break
                elif row_index == 3:
                    if data.value is None:
                        break
                    field_type_list.append(data.value)
                elif row_index == 4:
                    if data.value is None:
                        break
                    field_name_list.append(data.value)
                elif row_index >= 5:
                    # insert data
                    if data.value is None:
                        break

                    if col_index > 0:
                        q += ", "

                    q += '"{}"'.format(str(data.value))
                    empty_check = empty_check + 1

                # print(str(data.value), end=', ')

                col_index = col_index + 1

            if row_index == 4:
                # create table
                q = "create table {} (".format(sheet.title)

                col_index = 0
                for i in range(0, len(field_type_list)):
                    if i != 0:
                        q += ", "
                    field_type = field_type_list[i]
                    field_name = field_name_list[i]
                    q += "{} {}".format(field_name, field_type)
                    if primary_index == col_index:
                        q += " primary key"
                    col_index = col_index + 1

                q += ")"

                execute_query(conn, q)

                # generate cs file
                context = get_class_declare(sheet.title, primary_index, field_type_list, field_name_list)
                render_result = pystache.render(class_template, context)
                with open(args.cspath + "/class_{}.cs".format(sheet.title), "w") as class_render:
                    class_render.write(render_result)

            if row_index >= 5:
                if empty_check == 0:
                    break

                q += ")"
                # print(q)
                execute_query(conn, q)

            row_index = row_index + 1

        # print('')

    conn.commit()

    # cleanup
    cursor.close()
    conn.close()

    shutil.copy("master.db", args.output)

    # # lazy encrypt show
    # with open("master.db", mode="rb") as db:
    #     with open("master_converted", mode="wb") as convertedDB:
    #         db_data = db.read()
    #         convertedDB.write(base64.b64encode(db_data))
    #
    # # test for decrypt
    # with open("master_converted", mode="rb") as convertedDB:
    #     with open("master_recover.db", mode="wb") as db:
    #         db.write(base64.b64decode(convertedDB.read()))


if __name__ == '__main__':
    main()
